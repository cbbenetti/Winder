from pathlib import Path

from app.models.project import Project


def export_cable_excel(project: Project, path: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cable Schedule"
    headers = ["ID", "Label", "Type", "Signal Type", "From", "To", "Length (m)", "Notes"]
    ws.append(headers)
    for cable in sorted(project.cables, key=lambda c: c.id):
        ws.append([
            cable.id, cable.label, cable.cable_type, cable.signal_type,
            cable.from_endpoint, cable.to_endpoint, cable.length_m, cable.notes
        ])
    _auto_col_width(ws)
    wb.save(path)


def export_cable_csv(project: Project, path: str) -> None:
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ID", "Label", "Type", "Signal Type", "From", "To", "Length (m)", "Notes"])
        for cable in sorted(project.cables, key=lambda c: c.id):
            w.writerow([
                cable.id, cable.label, cable.cable_type, cable.signal_type,
                cable.from_endpoint, cable.to_endpoint, cable.length_m, cable.notes
            ])


def export_panel_excel(project: Project, path: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for panel in project.patch_panels:
        ws = wb.create_sheet(title=panel.id[:31])
        ws.append(["Port ID", "Row", "Col", "Label", "Signal Type", "Front Cable", "Rear Cable", "Notes"])
        for port in sorted(panel.ports, key=lambda p: (p.row, p.col)):
            ws.append([
                port.id, port.row, port.col, port.label, port.signal_type,
                port.front_cable_id, port.rear_cable_id, port.notes
            ])
        _auto_col_width(ws)
    if not wb.sheetnames:
        wb.create_sheet("No Panels")
    wb.save(path)


def export_daq_excel(project: Project, path: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DAQ Channels"
    ws.append(["Crate", "Crate Type", "Slot #", "Module Type", "Model", "Ch #", "Cable ID", "Signal Label", "Notes"])
    for crate in project.crates:
        for slot in crate.slots:
            for ch in slot.channels:
                ws.append([
                    crate.name or crate.id, crate.crate_type,
                    slot.slot_number, slot.module_type, slot.model,
                    ch.channel_number, ch.cable_id, ch.signal_label, ch.notes
                ])
    _auto_col_width(ws)
    wb.save(path)


def export_pdf(project: Project, path: str) -> None:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )

    doc = SimpleDocTemplate(str(path), pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = []

    # Cover
    story.append(Paragraph(f"System Report: {project.name}", styles["Title"]))
    story.append(Paragraph(f"Created: {project.created}", styles["Normal"]))
    story.append(Spacer(1, 24))

    # Cable schedule
    story.append(Paragraph("Cable Schedule", styles["Heading1"]))
    cable_data = [["ID", "Label", "Type", "Signal", "From", "To", "Length (m)", "Notes"]]
    for c in sorted(project.cables, key=lambda x: x.id):
        cable_data.append([c.id, c.label, c.cable_type, c.signal_type,
                           c.from_endpoint, c.to_endpoint, str(c.length_m), c.notes])
    story.append(_pdf_table(cable_data))
    story.append(PageBreak())

    # Patch panels
    story.append(Paragraph("Patch Panel Schedule", styles["Heading1"]))
    for panel in project.patch_panels:
        story.append(Paragraph(f"{panel.id} — {panel.name}", styles["Heading2"]))
        port_data = [["Port", "Label", "Signal", "Front Cable", "Rear Cable", "Notes"]]
        for port in sorted(panel.ports, key=lambda p: (p.row, p.col)):
            port_data.append([port.id, port.label, port.signal_type,
                               port.front_cable_id, port.rear_cable_id, port.notes])
        story.append(_pdf_table(port_data))
        story.append(Spacer(1, 12))
    story.append(PageBreak())

    # DAQ channel list
    story.append(Paragraph("DAQ Channel List", styles["Heading1"]))
    daq_data = [["Crate", "Type", "Slot", "Module", "Model", "Ch", "Cable ID", "Signal", "Notes"]]
    for crate in project.crates:
        for slot in crate.slots:
            for ch in slot.channels:
                daq_data.append([
                    crate.name or crate.id, crate.crate_type,
                    str(slot.slot_number), slot.module_type, slot.model,
                    str(ch.channel_number), ch.cable_id, ch.signal_label, ch.notes
                ])
    story.append(_pdf_table(daq_data))

    doc.build(story)


def _pdf_table(data):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d6a9f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eef2f7")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return t


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
